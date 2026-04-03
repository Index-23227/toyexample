"""
법인 매출 통합 (Task 2)
- 입력: data/법인_*.xlsx (5개)
- 출력: data/법인매출통합.xlsx (3개 시트)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import glob

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_FILE = os.path.join(DATA_DIR, "법인매출통합.xlsx")

# 스타일
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NUM_FORMAT = '#,##0'


def style_header(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def style_cell(cell, is_number=False):
    cell.border = THIN_BORDER
    if is_number:
        cell.number_format = NUM_FORMAT
        cell.alignment = Alignment(horizontal="right")


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


# ── 1. 법인 파일 읽기 ──
def read_corp_files():
    pattern = os.path.join(DATA_DIR, "법인_*.xlsx")
    files = sorted(glob.glob(pattern))
    all_data = []

    for fpath in files:
        fname = os.path.basename(fpath)
        # 법인_{코드}_{국가명}.xlsx 파싱
        m = re.match(r"법인_([A-Z0-9]+)_(.+)\.xlsx", fname)
        if not m:
            continue
        corp_code, corp_name = m.group(1), m.group(2)

        wb = openpyxl.load_workbook(fpath)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_data.append({
                "법인코드": corp_code,
                "법인명": corp_name,
                "월": row[0],
                "계정과목": row[1],
                "통화": row[2],
                "금액": row[3],
                "비고": row[4],
            })
        print(f"  {fname} → {corp_code} / {corp_name} ({ws.max_row - 1}건)")

    # 정렬: 법인코드 → 월 오름차순
    all_data.sort(key=lambda x: (x["법인코드"], x["월"]))
    return all_data


# ── 2. 집계 ──
def build_corp_summary(data):
    """법인별 매출합계 (현지통화 기준), 가나다순"""
    totals = {}
    for r in data:
        totals[r["법인명"]] = totals.get(r["법인명"], 0) + r["금액"]
    return dict(sorted(totals.items()))


def build_monthly_pivot(data):
    """법인명 × 월 교차표, 가나다순"""
    months = sorted(set(r["월"] for r in data))
    corps = sorted(set(r["법인명"] for r in data))

    pivot = {c: {m: 0 for m in months} for c in corps}
    for r in data:
        pivot[r["법인명"]][r["월"]] += r["금액"]

    return corps, months, pivot


# ── 3. Excel 출력 ──
def write_integrated_sheet(wb, data):
    """시트 1: 통합데이터"""
    ws = wb.create_sheet("통합데이터")
    headers = ["법인코드", "법인명", "월", "계정과목", "통화", "금액", "비고"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for r in data:
        ws.append([r["법인코드"], r["법인명"], r["월"], r["계정과목"], r["통화"], r["금액"], r["비고"]])
        row_num = ws.max_row
        for c in range(1, 8):
            style_cell(ws.cell(row_num, c), is_number=(c == 6))

    auto_width(ws)


def write_corp_summary_sheet(wb, summary):
    """시트 2: 법인별합계"""
    ws = wb.create_sheet("법인별합계")
    headers = ["법인명", "매출합계"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for name, total in summary.items():
        ws.append([name, total])
        r = ws.max_row
        style_cell(ws.cell(r, 1))
        style_cell(ws.cell(r, 2), is_number=True)

    auto_width(ws)


def write_pivot_sheet(wb, corps, months, pivot):
    """시트 3: 월별피벗"""
    ws = wb.create_sheet("월별피벗")
    headers = ["법인명"] + months + ["합계"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for corp in corps:
        row_vals = [pivot[corp][m] for m in months]
        row_total = sum(row_vals)
        ws.append([corp] + row_vals + [row_total])
        r = ws.max_row
        style_cell(ws.cell(r, 1))
        for c in range(2, len(headers) + 1):
            style_cell(ws.cell(r, c), is_number=True)

    auto_width(ws)


# ── 메인 ──
if __name__ == "__main__":
    print("[읽기] 법인 파일 스캔...")
    data = read_corp_files()
    print(f"[통합] 총 {len(data)}건\n")

    summary = build_corp_summary(data)
    corps, months, pivot = build_monthly_pivot(data)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_integrated_sheet(wb, data)
    write_corp_summary_sheet(wb, summary)
    write_pivot_sheet(wb, corps, months, pivot)

    wb.save(OUTPUT_FILE)
    print(f"[완료] {OUTPUT_FILE}\n")

    # 요약 출력
    print("── 법인별합계 (현지통화) ──")
    for name, total in summary.items():
        print(f"  {name:<6} {total:>15,}")

    print(f"\n── 월별피벗 ──")
    header = f"{'법인명':<6}" + "".join(f"{m:>14}" for m in months) + f"{'합계':>14}"
    print(header)
    print("-" * len(header))
    for corp in corps:
        vals = [pivot[corp][m] for m in months]
        line = f"{corp:<6}" + "".join(f"{v:>14,}" for v in vals) + f"{sum(vals):>14,}"
        print(line)
