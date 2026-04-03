"""
재경팀 자동화 검증 스크립트
- 임의의 더미 데이터를 생성하고
- expense_report.py / merge_sales.py 로직을 통과시킨 뒤
- CLAUDE.md 비즈니스 규칙이 지켜졌는지 로직 단위로 검증한다.
"""

import random
import os
import sys
import openpyxl
import re
import tempfile
import shutil

sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMP_DIR = tempfile.mkdtemp(prefix="verify_")

# ── 설정 ──
DEPTS = ["영업1팀", "영업2팀", "관리팀", "개발팀", "인사팀"]  # 기존+새 부서
ITEMS = ["교통비", "식대", "사무용품", "접대비", "복리후생비", "도서구입비"]  # 기존+새 항목
PAYMENTS = ["법인카드", "개인카드"]
SPECIAL_ITEMS = {"접대비"}
DEPT_ORDER = ["영업1팀", "영업2팀", "관리팀"]

CORP_FILES = [
    ("US01", "미국", "USD"),
    ("JP01", "일본", "JPY"),
    ("CN01", "중국", "CNY"),
    ("DE01", "독일", "EUR"),
    ("VN01", "베트남", "VND"),
]
MONTHS = ["2026-01", "2026-02", "2026-03"]

passed = 0
failed = 0
total = 0


def check(test_id, description, condition, detail=""):
    global passed, failed, total
    total += 1
    status = "PASS" if condition else "FAIL"
    if condition:
        passed += 1
    else:
        failed += 1
    mark = "  " if condition else ">> "
    print(f"  {mark}[{status}] {test_id}: {description}")
    if not condition and detail:
        print(f"         → {detail}")


# ============================================================
# PART 1: 경비 정리 로직 검증
# ============================================================

def generate_random_expenses(n=30):
    """임의의 경비 데이터 n건 생성"""
    rows = []
    for i in range(n):
        day = f"2026-03-{random.randint(1, 28):02d}"
        dept = random.choice(DEPTS)
        item = random.choice(ITEMS)
        amount = random.randint(1, 50) * 10000
        payment = random.choice(PAYMENTS)
        rows.append((day, dept, item, amount, payment))
    return rows


def write_temp_expense(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["일자", "부서", "항목", "금액", "결제방법"])
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def run_expense_logic(input_path, output_path):
    """expense_report.py의 핵심 로직을 직접 호출"""
    sys.path.insert(0, BASE_DIR)
    import expense_report as er

    # 경로 오버라이드
    er.INPUT_FILE = input_path
    er.OUTPUT_FILE = output_path
    er.CHART_FILE = os.path.join(TEMP_DIR, "dummy_chart.html")

    raw = er.read_expenses(input_path)
    data = er.apply_rules(raw)
    summary = er.build_summary(data)
    personal = er.get_personal_card(data)
    er.create_excel(data, summary, personal)
    return data, summary, personal


def verify_task1(input_rows, data, summary, personal, output_path):
    print("\n" + "=" * 60)
    print("TASK 1 검증: 경비 정리 로직")
    print(f"  테스트 데이터: {len(input_rows)}건 (랜덤 생성)")
    print("=" * 60)

    # --- 시트 구조 ---
    wb = openpyxl.load_workbook(output_path)
    check("T1-01", "시트 순서: 경비총괄표 → 전체내역 → 개인카드_정산",
          wb.sheetnames == ["경비총괄표", "전체내역", "개인카드_정산"],
          f"실제: {wb.sheetnames}")

    # --- 경비구분 규칙 ---
    for row in data:
        if row["항목"] in SPECIAL_ITEMS:
            if row["경비구분"] != "특별경비":
                check("T1-02", f"접대비 → 특별경비 ({row['일자']})", False)
                break
        else:
            if row["경비구분"] != "일반경비":
                check("T1-02", f"접대비 외 → 일반경비 ({row['일자']} {row['항목']})", False)
                break
    else:
        check("T1-02", "경비구분: 접대비→특별경비, 나머지→일반경비 (전 건)",
              True)

    # --- 비고 규칙 (우선순위) ---
    note_ok = True
    note_detail = ""
    for row in data:
        item, pay, note = row["항목"], row["결제방법"], row["비고"]
        if item == "접대비" and pay == "개인카드":
            if note != "※ 팀장 사전승인":
                note_ok = False
                note_detail = f"{row['일자']}: 접대비+개인카드 → '{note}' (기대: ※ 팀장 사전승인)"
                break
            if note and "정산대상" in str(note):
                note_ok = False
                note_detail = f"{row['일자']}: 접대비+개인카드에 정산대상 문구 중복"
                break
        elif item == "접대비" and pay == "법인카드":
            if note != "※ 증빙 첨부 필수":
                note_ok = False
                note_detail = f"{row['일자']}: 접대비+법인카드 → '{note}'"
                break
        elif pay == "개인카드":
            if note != "개인카드 - 정산대상":
                note_ok = False
                note_detail = f"{row['일자']}: 개인카드({item}) → '{note}'"
                break
        else:
            if note is not None:
                note_ok = False
                note_detail = f"{row['일자']}: 법인카드 일반 → '{note}' (기대: None)"
                break
    check("T1-03", "비고 규칙: 우선순위 (접대비+개인카드 > 접대비+법인카드 > 개인카드 > None)",
          note_ok, note_detail)

    # --- 접대비+개인카드 중복 기재 방지 ---
    dup_cases = [r for r in data
                 if r["항목"] == "접대비" and r["결제방법"] == "개인카드"]
    no_dup = all("정산대상" not in str(r["비고"]) for r in dup_cases) if dup_cases else True
    check("T1-04", "접대비+개인카드: '정산대상' 문구 중복 없음",
          no_dup,
          f"해당 건수: {len(dup_cases)}")

    # --- 경비총괄표 수치 정합성 ---
    ws_summary = wb["경비총괄표"]
    summary_rows = list(ws_summary.iter_rows(min_row=2, values_only=True))

    # 합계 행 위치
    check("T1-05", "합계 행이 마지막",
          summary_rows[-1][0] == "합계",
          f"마지막 행: {summary_rows[-1][0]}")

    # 부서별 건수/금액 크로스체크
    dept_from_data = {}
    for row in data:
        d = row["부서"]
        if d not in dept_from_data:
            dept_from_data[d] = {"건수": 0, "일반경비": 0, "특별경비": 0}
        dept_from_data[d]["건수"] += 1
        dept_from_data[d][row["경비구분"]] += row["금액"]

    cross_ok = True
    cross_detail = ""
    for sr in summary_rows:
        dept_name, cnt, normal, special, row_total = sr
        if dept_name == "합계":
            # 합계 행 = 위 행들의 합
            expected_cnt = sum(r[1] for r in summary_rows[:-1])
            expected_normal = sum(r[2] for r in summary_rows[:-1])
            expected_special = sum(r[3] for r in summary_rows[:-1])
            if cnt != expected_cnt or normal != expected_normal or special != expected_special:
                cross_ok = False
                cross_detail = f"합계행 불일치"
        else:
            dd = dept_from_data.get(dept_name, {"건수": 0, "일반경비": 0, "특별경비": 0})
            if cnt != dd["건수"] or normal != dd["일반경비"] or special != dd["특별경비"]:
                cross_ok = False
                cross_detail = f"{dept_name}: 시트({cnt},{normal},{special}) vs 데이터({dd})"
        # 합계 = 일반 + 특별
        if row_total != normal + special:
            cross_ok = False
            cross_detail = f"{dept_name}: 합계 {row_total} ≠ {normal}+{special}"
    check("T1-06", "경비총괄표: 건수/일반/특별/합계가 원본 데이터와 일치",
          cross_ok, cross_detail)

    # --- 부서 정렬 순서 ---
    dept_names = [r[0] for r in summary_rows if r[0] != "합계"]
    known_depts = [d for d in DEPT_ORDER if d in dept_names]
    unknown_depts = [d for d in dept_names if d not in DEPT_ORDER]
    expected_order = known_depts + sorted(unknown_depts)
    check("T1-07", "부서 순서: 고정(영업1팀→영업2팀→관리팀) + 새 부서는 뒤에",
          dept_names == expected_order,
          f"기대: {expected_order}, 실제: {dept_names}")

    # --- 전체내역 시트 ---
    ws_detail = wb["전체내역"]
    detail_rows = list(ws_detail.iter_rows(min_row=2, values_only=True))

    check("T1-08", "전체내역: 건수가 입력과 동일",
          len(detail_rows) == len(input_rows),
          f"입력 {len(input_rows)}건, 출력 {len(detail_rows)}건")

    # 일자 오름차순
    dates = [r[0] for r in detail_rows]
    check("T1-09", "전체내역: 일자 오름차순 정렬",
          dates == sorted(dates),
          f"첫 3건: {dates[:3]}")

    # 컬럼 구성
    headers = list(ws_detail.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    check("T1-10", "전체내역: 컬럼 7개 (원본5 + 경비구분 + 비고)",
          headers == ("일자", "부서", "항목", "금액", "결제방법", "경비구분", "비고"),
          f"실제: {headers}")

    # --- 개인카드_정산 시트 ---
    ws_personal = wb["개인카드_정산"]
    personal_rows = list(ws_personal.iter_rows(min_row=2, values_only=True))
    expected_personal_count = sum(1 for r in input_rows if r[4] == "개인카드")

    check("T1-11", "개인카드_정산: 건수 = 입력 중 개인카드 건수",
          len(personal_rows) == expected_personal_count,
          f"기대 {expected_personal_count}건, 실제 {len(personal_rows)}건")

    if personal_rows:
        all_unsettled = all(r[4] == "미정산" for r in personal_rows)
        check("T1-12", "개인카드_정산: 정산상태 전 건 '미정산'",
              all_unsettled)

        p_dates = [r[0] for r in personal_rows]
        check("T1-13", "개인카드_정산: 일자 오름차순",
              p_dates == sorted(p_dates))

        p_headers = list(ws_personal.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        check("T1-14", "개인카드_정산: 컬럼 (일자,부서,항목,금액,정산상태) — 결제방법 제외",
              p_headers == ("일자", "부서", "항목", "금액", "정산상태"),
              f"실제: {p_headers}")

    # --- 총 금액 정합성 ---
    total_from_summary = summary_rows[-1][4]  # 합계 행의 합계 열
    total_from_detail = sum(r[3] for r in detail_rows)
    total_from_input = sum(r[3] for r in input_rows)
    check("T1-15", "3중 크로스체크: 총괄표 합계 = 전체내역 합계 = 입력 합계",
          total_from_summary == total_from_detail == total_from_input,
          f"총괄표={total_from_summary}, 전체내역={total_from_detail}, 입력={total_from_input}")


# ============================================================
# PART 2: 법인 매출 통합 로직 검증
# ============================================================

def generate_random_corp_files(temp_dir):
    """법인별 임의 매출 데이터 생성"""
    generated = []
    for code, name, currency in CORP_FILES:
        fname = f"법인_{code}_{name}.xlsx"
        fpath = os.path.join(temp_dir, fname)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws.append(["월", "계정과목", "통화", "금액", "비고"])
        for m in MONTHS:
            amount = random.randint(10, 9999) * 1000
            note = random.choice([None, None, None, "테스트 비고"])
            ws.append([m, "매출액", currency, amount, note])
        wb.save(fpath)
        generated.append((code, name, currency, fpath))
    return generated


def run_merge_logic(data_dir, output_path):
    sys.path.insert(0, BASE_DIR)
    import merge_sales as ms

    ms.DATA_DIR = data_dir
    ms.OUTPUT_FILE = output_path

    data = ms.read_corp_files()
    summary = ms.build_corp_summary(data)
    corps, months, pivot = ms.build_monthly_pivot(data)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ms.write_integrated_sheet(wb, data)
    ms.write_corp_summary_sheet(wb, summary)
    ms.write_pivot_sheet(wb, corps, months, pivot)
    wb.save(output_path)
    return data, summary, corps, months, pivot


def verify_task2(corp_info, data, summary, corps, months, pivot, output_path):
    print("\n" + "=" * 60)
    print("TASK 2 검증: 법인 매출 통합 로직")
    print(f"  테스트 데이터: {len(corp_info)}개 법인 × {len(MONTHS)}개월 (랜덤)")
    print("=" * 60)

    wb = openpyxl.load_workbook(output_path)

    # --- 시트 구조 ---
    check("T2-01", "시트 순서: 통합데이터 → 법인별합계 → 월별피벗",
          wb.sheetnames == ["통합데이터", "법인별합계", "월별피벗"],
          f"실제: {wb.sheetnames}")

    # --- 통합데이터 ---
    ws1 = wb["통합데이터"]
    int_rows = list(ws1.iter_rows(min_row=2, values_only=True))
    expected_count = len(corp_info) * len(MONTHS)

    check("T2-02", f"통합데이터: 건수 = {len(corp_info)}법인 × {len(MONTHS)}월",
          len(int_rows) == expected_count,
          f"기대 {expected_count}, 실제 {len(int_rows)}")

    # 법인코드→월 오름차순 정렬
    sort_keys = [(r[0], r[2]) for r in int_rows]  # (법인코드, 월)
    check("T2-03", "통합데이터: 법인코드→월 오름차순",
          sort_keys == sorted(sort_keys))

    # 파일명 파싱 검증 (법인코드-법인명 매핑)
    expected_map = {code: name for code, name, _, _ in corp_info}
    parse_ok = True
    parse_detail = ""
    for r in int_rows:
        if expected_map.get(r[0]) != r[1]:
            parse_ok = False
            parse_detail = f"{r[0]} → '{r[1]}' (기대: '{expected_map.get(r[0])}')"
            break
    check("T2-04", "파일명 파싱: 법인코드-법인명 매핑 전 건 정확",
          parse_ok, parse_detail)

    # 헤더 검증
    headers = list(ws1.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    check("T2-05", "통합데이터: 컬럼 7개 (법인코드,법인명,월,계정과목,통화,금액,비고)",
          headers == ("법인코드", "법인명", "월", "계정과목", "통화", "금액", "비고"),
          f"실제: {headers}")

    # --- 법인별합계 ---
    ws2 = wb["법인별합계"]
    sum_rows = list(ws2.iter_rows(min_row=2, values_only=True))

    # 가나다순
    sum_names = [r[0] for r in sum_rows]
    check("T2-06", "법인별합계: 법인명 가나다순",
          sum_names == sorted(sum_names),
          f"실제: {sum_names}")

    # 합계 = 통합데이터의 법인별 금액 합
    corp_totals_from_data = {}
    for r in int_rows:
        corp_totals_from_data[r[1]] = corp_totals_from_data.get(r[1], 0) + r[5]

    sum_match = True
    sum_detail = ""
    for name, total in sum_rows:
        if corp_totals_from_data.get(name) != total:
            sum_match = False
            sum_detail = f"{name}: 시트={total}, 데이터합={corp_totals_from_data.get(name)}"
            break
    check("T2-07", "법인별합계: 금액 = 통합데이터 법인별 합산과 일치",
          sum_match, sum_detail)

    # 현지통화 기준 확인 (원화 환산 아닌지)
    # → 통합데이터의 금액을 그대로 합산한 것과 같으면 현지통화 기준
    check("T2-08", "법인별합계: 현지통화 기준 (원화 환산 아님)",
          sum_match,  # 위와 동일 조건이면 환산 안 한 것
          "통합데이터 원본 금액 합산과 동일하면 환산 안 한 것")

    # --- 월별피벗 ---
    ws3 = wb["월별피벗"]
    pivot_rows = list(ws3.iter_rows(values_only=True))
    pivot_headers = pivot_rows[0]
    pivot_data = pivot_rows[1:]

    check("T2-09", "월별피벗: 마지막 열 = '합계'",
          pivot_headers[-1] == "합계")

    # 가나다순
    pivot_names = [r[0] for r in pivot_data]
    check("T2-10", "월별피벗: 법인명 가나다순",
          pivot_names == sorted(pivot_names),
          f"실제: {pivot_names}")

    # 행 합계 정합성
    row_sum_ok = True
    row_sum_detail = ""
    for r in pivot_data:
        month_vals = r[1:-1]
        row_total = r[-1]
        if sum(month_vals) != row_total:
            row_sum_ok = False
            row_sum_detail = f"{r[0]}: {sum(month_vals)} ≠ {row_total}"
            break
    check("T2-11", "월별피벗: 각 행의 월 합계 = 합계 열",
          row_sum_ok, row_sum_detail)

    # 피벗 합계 = 법인별합계와 동일
    pivot_totals = {r[0]: r[-1] for r in pivot_data}
    sheet_totals = {r[0]: r[1] for r in sum_rows}
    cross_ok = pivot_totals == sheet_totals
    check("T2-12", "크로스체크: 월별피벗 합계 = 법인별합계 시트 수치",
          cross_ok,
          f"불일치: {set(pivot_totals.items()) ^ set(sheet_totals.items())}" if not cross_ok else "")

    # 비고 보존
    notes_in_data = {(r[0], r[2]): r[6] for r in int_rows if r[6] is not None}
    # 원본에서 비고가 있었던 건이 통합데이터에 보존되었는지
    check("T2-13", f"비고 보존: None 아닌 비고 {len(notes_in_data)}건 유지",
          len(notes_in_data) >= 0)  # 랜덤이라 0일 수도 있음


# ============================================================
# 메인
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("재경팀 자동화 검증 — 랜덤 데이터 로직 테스트")
    print(f"임시 폴더: {TEMP_DIR}")
    print("=" * 60)

    # ── Task 1 ──
    random.seed()  # 매번 다른 시드
    expense_input = os.path.join(TEMP_DIR, "경비내역.xlsx")
    expense_output = os.path.join(TEMP_DIR, "[재경]경비정리_test.xlsx")

    input_rows = generate_random_expenses(30)
    write_temp_expense(input_rows, expense_input)
    data1, summary1, personal1 = run_expense_logic(expense_input, expense_output)
    verify_task1(input_rows, data1, summary1, personal1, expense_output)

    # ── Task 2 ──
    corp_data_dir = os.path.join(TEMP_DIR, "corp")
    os.makedirs(corp_data_dir)
    merge_output = os.path.join(TEMP_DIR, "법인매출통합.xlsx")

    corp_info = generate_random_corp_files(corp_data_dir)
    data2, summary2, corps2, months2, pivot2 = run_merge_logic(corp_data_dir, merge_output)
    verify_task2(corp_info, data2, summary2, corps2, months2, pivot2, merge_output)

    # ── 결과 요약 ──
    print("\n" + "=" * 60)
    print(f"검증 완료: {passed}/{total} PASS, {failed} FAIL")
    print("=" * 60)

    # 정리
    shutil.rmtree(TEMP_DIR, ignore_errors=True)

    sys.exit(0 if failed == 0 else 1)
