"""
경비 정리 자동화 + 시각화
- 입력: data/경비내역.xlsx
- 출력: data/[재경]경비정리_202603_v1.xlsx (3개 시트)
- 시각화: data/경비_시각화.html (인터랙티브 차트)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict
import os

# ── 설정 ──
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
INPUT_FILE = os.path.join(DATA_DIR, "경비내역.xlsx")
OUTPUT_FILE = os.path.join(DATA_DIR, "[재경]경비정리_202603_v1.xlsx")
CHART_FILE = os.path.join(DATA_DIR, "경비_시각화.html")

# 부서 정렬 순서 (사내 고정)
DEPT_ORDER = ["영업1팀", "영업2팀", "관리팀"]

# 특별경비 항목
SPECIAL_EXPENSE_ITEMS = {"접대비"}


# ── 1. 데이터 읽기 ──
def read_expenses(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 헤더 제외
    data = []
    for row in rows:
        data.append({
            "일자": row[0],
            "부서": row[1],
            "항목": row[2],
            "금액": row[3],
            "결제방법": row[4],
        })
    return data


# ── 2. 비즈니스 규칙 적용 ──
def classify_expense(item):
    """경비 구분: 접대비 → 특별경비, 나머지 → 일반경비"""
    return "특별경비" if item in SPECIAL_EXPENSE_ITEMS else "일반경비"


def generate_note(item, payment):
    """비고 자동 생성 (우선순위: 접대비+개인카드 > 접대비+법인카드 > 개인카드)"""
    if item == "접대비" and payment == "개인카드":
        return "※ 팀장 사전승인"
    if item == "접대비" and payment == "법인카드":
        return "※ 증빙 첨부 필수"
    if payment == "개인카드":
        return "개인카드 - 정산대상"
    return None


def apply_rules(data):
    for row in data:
        row["경비구분"] = classify_expense(row["항목"])
        row["비고"] = generate_note(row["항목"], row["결제방법"])
    return data


# ── 3. 집계 ──
def build_summary(data):
    """부서별 건수/일반경비/특별경비 집계"""
    raw = {}
    for row in data:
        dept = row["부서"]
        if dept not in raw:
            raw[dept] = {"건수": 0, "일반경비": 0, "특별경비": 0}
        raw[dept]["건수"] += 1
        raw[dept][row["경비구분"]] += row["금액"]

    # 정렬: 고정 순서(DEPT_ORDER) 먼저, 나머지는 가나다순
    summary = OrderedDict()
    for dept in DEPT_ORDER:
        if dept in raw:
            summary[dept] = raw.pop(dept)
    for dept in sorted(raw.keys()):
        summary[dept] = raw[dept]

    return summary


def get_personal_card(data):
    """개인카드 결제건 추출"""
    return [r for r in data if r["결제방법"] == "개인카드"]


# ── 4. Excel 출력 ──
# 스타일 정의
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NUM_FORMAT = '#,##0'


def style_header(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def style_cell(cell, is_number=False):
    cell.border = THIN_BORDER
    if is_number:
        cell.number_format = NUM_FORMAT
        cell.alignment = Alignment(horizontal="right")


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def write_summary_sheet(wb, summary):
    """시트 1: 경비총괄표"""
    ws = wb.create_sheet("경비총괄표")
    headers = ["구분", "건수", "일반경비", "특별경비", "합계"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    total = {"건수": 0, "일반경비": 0, "특별경비": 0}

    for dept, vals in summary.items():
        row_sum = vals["일반경비"] + vals["특별경비"]
        row_data = [dept, vals["건수"], vals["일반경비"], vals["특별경비"], row_sum]
        ws.append(row_data)
        r = ws.max_row
        style_cell(ws.cell(r, 1))
        for c in range(2, 6):
            style_cell(ws.cell(r, c), is_number=True)
        total["건수"] += vals["건수"]
        total["일반경비"] += vals["일반경비"]
        total["특별경비"] += vals["특별경비"]

    # 합계 행
    total_sum = total["일반경비"] + total["특별경비"]
    ws.append(["합계", total["건수"], total["일반경비"], total["특별경비"], total_sum])
    r = ws.max_row
    for c in range(1, 6):
        cell = ws.cell(r, c)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        style_cell(cell, is_number=(c > 1))

    auto_width(ws)
    return total


def write_detail_sheet(wb, data):
    """시트 2: 전체내역"""
    ws = wb.create_sheet("전체내역")
    headers = ["일자", "부서", "항목", "금액", "결제방법", "경비구분", "비고"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    special_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for row in sorted(data, key=lambda x: x["일자"]):
        ws.append([
            row["일자"], row["부서"], row["항목"], row["금액"],
            row["결제방법"], row["경비구분"], row["비고"],
        ])
        r = ws.max_row
        for c in range(1, 8):
            cell = ws.cell(r, c)
            style_cell(cell, is_number=(c == 4))
            if row["경비구분"] == "특별경비":
                cell.fill = special_fill

    auto_width(ws)


def write_personal_sheet(wb, personal):
    """시트 3: 개인카드_정산"""
    ws = wb.create_sheet("개인카드_정산")
    headers = ["일자", "부서", "항목", "금액", "정산상태"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for row in sorted(personal, key=lambda x: x["일자"]):
        ws.append([row["일자"], row["부서"], row["항목"], row["금액"], "미정산"])
        r = ws.max_row
        for c in range(1, 6):
            style_cell(ws.cell(r, c), is_number=(c == 4))

    auto_width(ws)


def create_excel(data, summary, personal):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    total = write_summary_sheet(wb, summary)
    write_detail_sheet(wb, data)
    write_personal_sheet(wb, personal)

    wb.save(OUTPUT_FILE)
    print(f"[완료] {OUTPUT_FILE}")
    return total


# ── 5. 시각화 (Plotly HTML) ──
def create_visualization(data, summary):
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    # --- 데이터 준비 ---
    depts = list(summary.keys())
    normal = [summary[d]["일반경비"] for d in depts]
    special = [summary[d]["특별경비"] for d in depts]
    totals = [n + s for n, s in zip(normal, special)]

    # 항목별 집계
    item_totals = {}
    for row in data:
        item_totals[row["항목"]] = item_totals.get(row["항목"], 0) + row["금액"]

    # 결제방법별 집계
    pay_totals = {}
    for row in data:
        pay_totals[row["결제방법"]] = pay_totals.get(row["결제방법"], 0) + row["금액"]

    # 일별 추이
    daily = {}
    for row in data:
        daily[row["일자"]] = daily.get(row["일자"], 0) + row["금액"]
    daily_sorted = sorted(daily.items())

    # --- 차트 구성 (2×2) ---
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=(
            "부서별 경비 (일반 vs 특별)",
            "항목별 경비 비중",
            "결제방법별 비중",
            "일별 경비 추이",
        ),
        specs=[
            [{"type": "bar"}, {"type": "pie"}],
            [{"type": "pie"}, {"type": "scatter"}],
        ],
        vertical_spacing=0.15,
        horizontal_spacing=0.1,
    )

    colors_blue = ["#2F5496", "#5B9BD5", "#A5C8E1"]
    colors_orange = ["#ED7D31", "#F4B183"]

    # 차트 1: 부서별 스택 바
    fig.add_trace(go.Bar(
        name="일반경비", x=depts, y=normal,
        marker_color="#5B9BD5", text=[f"{v:,}" for v in normal], textposition="inside",
    ), row=1, col=1)
    fig.add_trace(go.Bar(
        name="특별경비", x=depts, y=special,
        marker_color="#ED7D31", text=[f"{v:,}" for v in special], textposition="inside",
    ), row=1, col=1)
    fig.update_layout(barmode="stack")

    # 차트 2: 항목별 파이
    fig.add_trace(go.Pie(
        labels=list(item_totals.keys()),
        values=list(item_totals.values()),
        marker_colors=["#5B9BD5", "#A5C8E1", "#2F5496", "#ED7D31"],
        textinfo="label+percent", hole=0.3,
    ), row=1, col=2)

    # 차트 3: 결제방법 파이
    fig.add_trace(go.Pie(
        labels=list(pay_totals.keys()),
        values=list(pay_totals.values()),
        marker_colors=["#5B9BD5", "#ED7D31"],
        textinfo="label+percent", hole=0.3,
    ), row=2, col=1)

    # 차트 4: 일별 추이
    fig.add_trace(go.Scatter(
        x=[d[0] for d in daily_sorted],
        y=[d[1] for d in daily_sorted],
        mode="lines+markers+text",
        text=[f"{d[1]:,}" for d in daily_sorted],
        textposition="top center", textfont=dict(size=9),
        line=dict(color="#2F5496", width=2),
        marker=dict(size=7),
    ), row=2, col=2)

    # --- 상단 KPI 카드 (annotation) ---
    total_amount = sum(totals)
    total_count = sum(summary[d]["건수"] for d in depts)
    personal_count = sum(1 for r in data if r["결제방법"] == "개인카드")
    personal_amount = sum(r["금액"] for r in data if r["결제방법"] == "개인카드")

    fig.update_layout(
        title=dict(
            text=(
                f"<b>재경팀 경비 정리 현황  |  2026년 3월</b><br>"
                f"<span style='font-size:13px; color:#666'>"
                f"총 {total_count}건 · ₩{total_amount:,}  |  "
                f"일반경비 ₩{sum(normal):,} · 특별경비 ₩{sum(special):,}  |  "
                f"개인카드 정산대상 {personal_count}건 · ₩{personal_amount:,}"
                f"</span>"
            ),
            x=0.5, font=dict(size=16),
        ),
        height=750,
        width=1100,
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.05, xanchor="center", x=0.5),
        template="plotly_white",
        font=dict(family="Malgun Gothic, sans-serif"),
    )

    fig.write_html(CHART_FILE, include_plotlyjs="cdn")
    print(f"[완료] {CHART_FILE}")


# ── 메인 ──
if __name__ == "__main__":
    # 1. 읽기
    raw = read_expenses(INPUT_FILE)
    print(f"[읽기] {len(raw)}건 로드")

    # 2. 규칙 적용
    data = apply_rules(raw)

    # 3. 집계
    summary = build_summary(data)
    personal = get_personal_card(data)

    # 4. Excel 생성
    create_excel(data, summary, personal)

    # 5. 시각화
    create_visualization(data, summary)

    # 6. 요약 출력
    print("\n── 경비총괄표 ──")
    print(f"{'구분':<8} {'건수':>4} {'일반경비':>10} {'특별경비':>10} {'합계':>10}")
    print("-" * 48)
    grand = {"건수": 0, "일반": 0, "특별": 0}
    for dept, v in summary.items():
        s = v["일반경비"] + v["특별경비"]
        print(f"{dept:<8} {v['건수']:>4} {v['일반경비']:>10,} {v['특별경비']:>10,} {s:>10,}")
        grand["건수"] += v["건수"]
        grand["일반"] += v["일반경비"]
        grand["특별"] += v["특별경비"]
    print("-" * 48)
    gt = grand["일반"] + grand["특별"]
    print(f"{'합계':<8} {grand['건수']:>4} {grand['일반']:>10,} {grand['특별']:>10,} {gt:>10,}")

    print(f"\n── 개인카드 정산대상: {len(personal)}건 ──")
    for r in personal:
        print(f"  {r['일자']}  {r['부서']}  {r['항목']}  {r['금액']:>8,}  미정산")
