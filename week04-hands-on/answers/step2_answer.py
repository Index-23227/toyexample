"""Step 2 테스트: 경비내역.xlsx → CLAUDE.md 하네스 양식대로 정리"""
import openpyxl
from openpyxl.styles import Font, numbers
from collections import defaultdict

# 경비내역.xlsx 읽기
wb = openpyxl.load_workbook("data/경비내역.xlsx")
ws = wb.active

headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append(dict(zip(headers, row)))

print(f"총 {len(rows)}건 읽음\n")

# === 사내 경비 규칙 ===
# 일반경비 = 교통비 + 식대 + 사무용품
# 특별경비 = 접대비
일반경비_항목 = {"교통비", "식대", "사무용품"}

def 경비구분(항목):
    return "일반경비" if 항목 in 일반경비_항목 else "특별경비"

def 비고_규칙(항목, 금액, 결제방법):
    """비고 표시 규칙 (위에서부터 우선 적용)"""
    if 항목 == "접대비" and 금액 >= 100000:
        return "※ 증빙 첨부 필수"
    if 항목 == "접대비" and 금액 >= 50000:
        return "※ 팀장 사전승인"
    if 결제방법 == "개인카드":
        return "개인카드 - 정산대상"
    return ""

# 부서 순서 (사내 규칙: 영업1팀→영업2팀→관리팀→합계)
부서순서 = ["영업1팀", "영업2팀", "관리팀"]

# === 시트별 데이터 계산 ===

# 시트1: 경비총괄표 — 부서별 건수, 일반경비, 특별경비, 합계
총괄 = {}
for dept in 부서순서:
    총괄[dept] = {"건수": 0, "일반경비": 0, "특별경비": 0}

for r in rows:
    dept = r["부서"]
    amt = r["금액"]
    구분 = 경비구분(r["항목"])
    총괄[dept]["건수"] += 1
    총괄[dept][구분] += amt

# === 엑셀 저장 ===
out_wb = openpyxl.Workbook()
bold = Font(bold=True)
comma_fmt = '#,##0'

# --- 시트1: 경비총괄표 ---
ws1 = out_wb.active
ws1.title = "경비총괄표"
ws1.append(["구분", "건수", "일반경비", "특별경비", "합계"])
# 헤더 굵게
for cell in ws1[1]:
    cell.font = bold

total_건수 = 0
total_일반 = 0
total_특별 = 0

for dept in 부서순서:
    d = 총괄[dept]
    합계 = d["일반경비"] + d["특별경비"]
    ws1.append([dept, d["건수"], d["일반경비"], d["특별경비"], 합계])
    total_건수 += d["건수"]
    total_일반 += d["일반경비"]
    total_특별 += d["특별경비"]

# 합계 행 (굵은 글씨)
합계행 = ws1.max_row + 1
ws1.append(["합계", total_건수, total_일반, total_특별, total_일반 + total_특별])
for cell in ws1[합계행]:
    cell.font = bold

# 금액 셀에 천 단위 콤마 서식
for row in ws1.iter_rows(min_row=2, min_col=3, max_col=5):
    for cell in row:
        cell.number_format = comma_fmt

# --- 시트2: 전체내역 ---
ws2 = out_wb.create_sheet("전체내역")
ws2.append(["일자", "부서", "항목", "금액", "결제방법", "경비구분", "비고"])
for cell in ws2[1]:
    cell.font = bold

# 일자 기준 오름차순 정렬
sorted_rows = sorted(rows, key=lambda r: r["일자"])
for r in sorted_rows:
    구분 = 경비구분(r["항목"])
    비고 = 비고_규칙(r["항목"], r["금액"], r["결제방법"])
    ws2.append([r["일자"], r["부서"], r["항목"], r["금액"], r["결제방법"], 구분, 비고])

for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.number_format = comma_fmt

# --- 시트3: 개인카드_정산 ---
ws3 = out_wb.create_sheet("개인카드_정산")
ws3.append(["일자", "부서", "항목", "금액", "정산상태"])
for cell in ws3[1]:
    cell.font = bold

personal = [r for r in sorted_rows if r["결제방법"] == "개인카드"]
for r in personal:
    ws3.append([r["일자"], r["부서"], r["항목"], r["금액"], "미정산"])

for row in ws3.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.number_format = comma_fmt

# 저장
output_file = "data/[재경]경비정리_202603_v1.xlsx"
out_wb.save(output_file)

# 터미널 출력
print("=== 경비총괄표 ===")
print(f"{'구분':<10} {'건수':>4} {'일반경비':>12} {'특별경비':>12} {'합계':>12}")
print("-" * 54)
for dept in 부서순서:
    d = 총괄[dept]
    합 = d["일반경비"] + d["특별경비"]
    print(f"{dept:<10} {d['건수']:>4} {d['일반경비']:>12,.0f} {d['특별경비']:>12,.0f} {합:>12,.0f}")
print(f"{'합계':<10} {total_건수:>4} {total_일반:>12,.0f} {total_특별:>12,.0f} {total_일반+total_특별:>12,.0f}")

print(f"\n개인카드 정산 대상: {len(personal)}건")
print(f"\n결과 저장: {output_file}")
