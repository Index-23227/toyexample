"""Step 3 테스트: 법인별 매출 파일 5개를 하나로 합치기"""
import openpyxl
import glob
import re

# 법인_ 으로 시작하는 파일 찾기
files = glob.glob("data/법인_*.xlsx")
print(f"찾은 파일: {len(files)}개")
for f in files:
    print(f"  - {f}")

# 모든 파일 읽어서 합치기
all_rows = []
for filepath in files:
    # 파일명에서 법인코드, 법인명 추출
    filename = filepath.replace("\\", "/").split("/")[-1]  # 윈도우 경로 처리
    match = re.search(r"법인_(\w+)_(.+)\.xlsx", filename)
    if not match:
        print(f"  ⚠ 파일명 파싱 실패: {filename}")
        continue
    corp_code = match.group(1)
    corp_name = match.group(2)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        row_dict["법인코드"] = corp_code
        row_dict["법인명"] = corp_name
        all_rows.append(row_dict)

print(f"\n총 {len(all_rows)}행 합쳐짐")

# 법인별 합계
from collections import defaultdict
corp_totals = defaultdict(float)
for r in all_rows:
    corp_totals[r["법인명"]] += r["금액"]

print("\n=== 법인별 매출 합계 ===")
for name, total in sorted(corp_totals.items()):
    print(f"  {name}: {total:,.0f}")

# 통합 엑셀 저장
out_wb = openpyxl.Workbook()

# 시트1: 전체 데이터
ws1 = out_wb.active
ws1.title = "통합데이터"
out_headers = ["법인코드", "법인명", "월", "계정과목", "통화", "금액", "비고"]
ws1.append(out_headers)
for r in all_rows:
    ws1.append([r.get(h, "") for h in out_headers])

# 시트2: 법인별 합계
ws2 = out_wb.create_sheet("법인별합계")
ws2.append(["법인명", "매출합계"])
for name, total in sorted(corp_totals.items()):
    ws2.append([name, total])

# 시트3: 월별 피벗 (법인×월 매출 합계)
pivot = defaultdict(lambda: defaultdict(float))
months = set()
for r in all_rows:
    pivot[r["법인명"]][r["월"]] += r["금액"]
    months.add(r["월"])

months_sorted = sorted(months)
ws3 = out_wb.create_sheet("월별피벗")
ws3.append(["법인명"] + months_sorted + ["합계"])
for name in sorted(pivot.keys()):
    row_total = sum(pivot[name][m] for m in months_sorted)
    ws3.append([name] + [pivot[name][m] for m in months_sorted] + [row_total])

print("\n=== 법인별 × 월별 매출 ===")
print(f"{'법인명':<8}", end="")
for m in months_sorted:
    print(f" {m:>14}", end="")
print(f" {'합계':>14}")
print("-" * (8 + 15 * (len(months_sorted) + 1)))
for name in sorted(pivot.keys()):
    print(f"{name:<8}", end="")
    for m in months_sorted:
        print(f" {pivot[name][m]:>14,.0f}", end="")
    row_total = sum(pivot[name][m] for m in months_sorted)
    print(f" {row_total:>14,.0f}")

out_wb.save("data/법인매출통합.xlsx")
print("\n결과 저장: data/법인매출통합.xlsx")
